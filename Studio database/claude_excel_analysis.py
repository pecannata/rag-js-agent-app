import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re

def claude_analyze_excel(file_path, sheet_name):
    """
    Claude's comprehensive analysis of the Excel file
    """
    print("=== CLAUDE'S EXCEL ANALYSIS ===")
    
    # Load with openpyxl for raw cell access
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    print(f"Analyzing sheet: {sheet_name}")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Get all non-empty cells
    all_data = {}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                all_data[(row, col)] = str(cell.value).strip()
    
    print(f"Found {len(all_data)} non-empty cells")
    
    # Print complete cell map for analysis
    print("\n=== COMPLETE CELL MAP ===")
    for row in range(1, min(15, ws.max_row + 1)):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            if (row, col) in all_data:
                content = all_data[(row, col)]
                if len(content) > 30:
                    content = content[:30] + "..."
                row_cells.append(f"C{col}:'{content}'")
        if row_cells:
            print(f"Row {row}: {', '.join(row_cells)}")
    
    # Identify structure
    print("\n=== STRUCTURE IDENTIFICATION ===")
    
    # Find header/dates/days rows
    date_row = None
    day_row = None
    time_start_row = None
    
    for row in range(1, 15):
        row_values = [all_data.get((row, col), '') for col in range(1, ws.max_column + 1)]
        
        # Check for dates (2024-xx-xx pattern)
        date_count = sum(1 for val in row_values if re.search(r'2024-\d{2}-\d{2}', val))
        if date_count >= 5:  # At least 5 dates
            date_row = row
            print(f"Date row found: {row}")
        
        # Check for day names
        day_names = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        day_count = sum(1 for val in row_values if val.upper() in day_names)
        if day_count >= 5:  # At least 5 days
            day_row = row
            print(f"Day row found: {row}")
        
        # Check for time patterns
        time_count = sum(1 for val in row_values if re.search(r'\d{1,2}:\d{2}', val))
        if time_count >= 1 and not time_start_row:
            # Make sure it's not the date row
            if row != date_row:
                time_start_row = row
                print(f"Time data starts at row: {row}")
    
    # Extract dates
    dates = []
    if date_row:
        for col in range(1, ws.max_column + 1):
            val = all_data.get((date_row, col), '')
            if '2024' in val:
                dates.append(val.split()[0])
    
    # Extract days  
    days = []
    day_columns = []
    if day_row:
        for col in range(1, ws.max_column + 1):
            val = all_data.get((day_row, col), '').upper()
            if val in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
                days.append(val)
                day_columns.append(col)
    
    print(f"Extracted {len(dates)} dates: {dates}")
    print(f"Extracted {len(days)} days: {days}")
    print(f"Day columns: {day_columns}")
    
    # Extract ALL time slots and lessons
    print("\n=== EXTRACTING ALL LESSON DATA ===")
    
    lesson_data = []
    
    # Look through all rows for time patterns
    for row in range(time_start_row or 5, ws.max_row + 1):
        row_values = [all_data.get((row, col), '') for col in range(1, ws.max_column + 1)]
        
        # Check if this row has any time information
        times_in_row = []
        for val in row_values:
            time_matches = re.findall(r'\d{1,2}:\d{2}(?::\d{2})?(?:\s*(?:AM|PM))?', val, re.IGNORECASE)
            times_in_row.extend(time_matches)
        
        # Check if row has student names
        student_names = ['GEMMA', 'EVERLY', 'REMI', 'LARKIN', 'MILA', 'REESE']
        has_students = any(any(name in val.upper() for name in student_names) for val in row_values)
        
        # Check if row has lesson content
        has_content = any(val.strip() and not re.match(r'^\d{4}-\d{2}-\d{2}', val) for val in row_values)
        
        if times_in_row or has_students or (has_content and row <= 15):
            print(f"Row {row}: Times={times_in_row}, Students={has_students}, Content={has_content}")
            print(f"  Data: {[f'C{i+1}:{v[:40]}' for i, v in enumerate(row_values) if v.strip()]}")
            
            # Extract time slot for this row
            primary_time = None
            for col in range(1, 3):  # Check first few columns for primary time
                val = all_data.get((row, col), '')
                time_match = re.search(r'\d{1,2}:\d{2}(?::\d{2})?', val)
                if time_match:
                    primary_time = time_match.group()
                    break
            
            if primary_time or has_students or has_content:
                lesson_row = [primary_time or '']
                
                # Extract data for each day column
                for col in day_columns:
                    cell_value = all_data.get((row, col), '')
                    lesson_row.append(cell_value)
                
                lesson_data.append(lesson_row)
    
    # Create CSV structure
    headers = ['Time'] + [f"{day}\n{date}" for day, date in zip(days, dates)]
    
    # Filter out empty rows
    lesson_data = [row for row in lesson_data if any(cell.strip() for cell in row)]
    
    return headers, lesson_data, all_data

# Run the analysis
if __name__ == "__main__":
    file_path = 'Private_lesson_Calendar.xlsx'
    sheet_name = '99-915'
    
    headers, lesson_data, raw_data = claude_analyze_excel(file_path, sheet_name)
    
    print(f"\n=== FINAL EXTRACTION RESULTS ===")
    print(f"Headers: {headers}")
    print(f"Extracted {len(lesson_data)} lesson rows")
    
    # Create DataFrame and save
    import pandas as pd
    df = pd.DataFrame(lesson_data, columns=headers)
    
    output_file = 'CLAUDE_COMPLETE_ANALYSIS.csv'
    df.to_csv(output_file, index=False)
    
    print(f"\nSaved to: {output_file}")
    print("\nFinal result:")
    print(df.to_string())
    
    # Show what we captured vs what exists
    print(f"\n=== DATA COMPLETENESS CHECK ===")
    print(f"Total non-empty cells in Excel: {len(raw_data)}")
    print(f"Lesson data rows captured: {len(lesson_data)}")
