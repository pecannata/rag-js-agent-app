from openpyxl import load_workbook
import re
from collections import defaultdict

def debug_colors_and_values():
    """Debug script to see all colors and their associated values"""
    
    file_path = 'Private lesson Calendar.xlsx'
    sheet_name = '69 -615'
    
    print(f"=== DEBUG: ANALYZING ALL COLORS AND VALUES ===")
    
    # Load workbook with formatting
    wb_format = load_workbook(file_path, data_only=False)
    ws_format = wb_format[sheet_name]
    
    # Load workbook with data
    wb_data = load_workbook(file_path, data_only=True)
    ws_data = wb_data[sheet_name]
    
    color_data = defaultdict(list)
    all_colors = set()
    
    print(f"Sheet dimensions: {ws_data.max_row} rows × {ws_data.max_column} columns")
    
    # Collect all cell data with colors
    for row in range(1, ws_data.max_row + 1):
        for col in range(1, ws_data.max_column + 1):
            cell_data_val = ws_data.cell(row=row, column=col)
            cell_format_val = ws_format.cell(row=row, column=col)
            
            if cell_data_val.value is not None:
                value = str(cell_data_val.value).strip()
                
                # Get cell fill color
                fill = cell_format_val.fill
                color = None
                
                if fill and fill.start_color and fill.start_color.rgb:
                    if fill.start_color.rgb != '00000000':  # Not default/no color
                        color = fill.start_color.rgb
                        all_colors.add(color)
                
                color_data[color].append({
                    'row': row,
                    'col': col,
                    'value': value
                })
    
    print(f"\nFound {len(all_colors)} unique colors (excluding no-color)")
    
    # Show all colors and their sample values
    print("\n=== ALL COLORS AND THEIR VALUES ===")
    for color in sorted(all_colors):
        print(f"\nColor: {color}")
        sample_values = [item['value'] for item in color_data[color][:10]]  # First 10 values
        print(f"  Sample values: {sample_values}")
        
        # Look for potential teacher names in this color
        teacher_candidates = []
        for item in color_data[color]:
            value = item['value']
            # Check if this looks like a teacher name (ALL CAPS, looks like a name)
            if (value.isupper() and 
                len(value) > 2 and 
                not re.search(r'\d{4}-\d{2}-\d{2}|DAILY|SCHEDULE|STUDIO|MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY|\d{2}:\d{2}', value)):
                teacher_candidates.append(value)
        
        if teacher_candidates:
            print(f"  Potential teachers: {list(set(teacher_candidates))}")
    
    # Show cells with no color but containing data
    print(f"\n=== CELLS WITH NO COLOR ===")
    no_color_values = [item['value'] for item in color_data[None][:20]]  # First 20
    print(f"Sample values with no color: {no_color_values}")
    
    return color_data, all_colors

if __name__ == "__main__":
    color_data, all_colors = debug_colors_and_values()
