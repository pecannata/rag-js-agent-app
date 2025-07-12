from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import re

def extract_teachers_by_color():
    """Extract cell colors from the 69-615 sheet to identify teachers"""
    
    file_path = 'Studio database/Private lesson Calendar.xlsx'
    sheet_name = '69 -615'
    
    # Load workbook WITHOUT data_only to preserve formatting
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    print(f"=== EXTRACTING TEACHERS BY COLOR: {sheet_name} ===")
    print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Get all cell data including colors
    cell_data = {}
    color_map = {}
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            if cell.value is not None:
                # Get cell value
                value = str(cell.value).strip()
                
                # Get cell fill color
                fill = cell.fill
                color = None
                
                if fill and fill.start_color and fill.start_color.rgb:
                    if fill.start_color.rgb != '00000000':  # Not default/no color
                        color = fill.start_color.rgb
                        
                cell_data[(row, col)] = {
                    'value': value,
                    'color': color
                }
                
                # Track unique colors
                if color and color not in color_map:
                    color_map[color] = []
                if color:
                    color_map[color].append((row, col, value))
    
    print(f"Found {len(cell_data)} cells with data")
    print(f"Found {len(color_map)} unique colors")
    
    # Analyze color patterns
    print("\\n=== COLOR ANALYSIS ===")
    for color, cells in color_map.items():
        print(f"Color {color}: {len(cells)} cells")
        # Show sample cells for this color
        sample_values = [cell[2] for cell in cells[:5]]  # First 5 values
        print(f"  Sample values: {sample_values}")
    
    # Find structure rows (same as before)
    date_row = 3
    day_row = 4  
    studio_row = 5
    time_start_row = 6
    
    # Extract day-studio columns
    day_studio_cols = []
    for col in range(1, ws.max_column + 1):
        day_val = cell_data.get((day_row, col), {}).get('value', '').upper()
        studio_val = cell_data.get((studio_row, col), {}).get('value', '')
        if day_val in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
            day_studio_cols.append((col, day_val, studio_val))
    
    print(f"\\nFound {len(day_studio_cols)} day-studio combinations")
    
    # Extract lesson data with colors (teachers)
    print("\\n=== EXTRACTING LESSONS WITH TEACHER COLORS ===")
    
    teacher_data = []
    lesson_data = []
    
    for row in range(time_start_row, ws.max_row + 1):
        # Get time
        time_cell = cell_data.get((row, 2), {})
        time_val = time_cell.get('value', '')
        
        if re.search(r'\d{1,2}:\d{2}:\d{2}', time_val):
            teacher_row = [time_val]
            lesson_row = [time_val]
            
            for col, day, studio in day_studio_cols:
                cell_info = cell_data.get((row, col), {})
                value = cell_info.get('value', '')
                color = cell_info.get('color', '')
                
                teacher_row.append(color if color else '')
                lesson_row.append(value)
            
            # Only add if there's data
            if any(cell.strip() for cell in teacher_row[1:]) or any(cell.strip() for cell in lesson_row[1:]):
                teacher_data.append(teacher_row)
                lesson_data.append(lesson_row)
                
                # Show color info for this time slot
                colored_cells = [(i, color) for i, color in enumerate(teacher_row[1:]) if color]
                if colored_cells:
                    print(f"Time {time_val}: {len(colored_cells)} colored cells")
    
    # Create headers
    headers = ['Time']
    for col, day, studio in day_studio_cols:
        date_val = cell_data.get((date_row, col), {}).get('value', '')
        if '2025' in str(date_val):
            date = str(date_val).split()[0]
        else:
            date = ''
        headers.append(f"{day} {studio} ({date})")
    
    # Save teacher colors CSV
    teacher_df = pd.DataFrame(teacher_data, columns=headers)
    teacher_output = 'Studio database/Private_lesson_Calendar_69-615_TEACHERS_BY_COLOR.csv'
    teacher_df.to_csv(teacher_output, index=False)
    
    # Save lesson data CSV  
    lesson_df = pd.DataFrame(lesson_data, columns=headers)
    lesson_output = 'Studio database/Private_lesson_Calendar_69-615_LESSONS_WITH_COLORS.csv'
    lesson_df.to_csv(lesson_output, index=False)
    
    print(f"\\n=== RESULTS ===")
    print(f"Teacher colors saved to: {teacher_output}")
    print(f"Lesson data saved to: {lesson_output}")
    print(f"Time slots: {len(teacher_data)}")
    
    # Show color mapping summary
    print("\\n=== COLOR TO TEACHER MAPPING ===")
    print("You'll need to manually map these colors to teacher names:")
    for i, (color, cells) in enumerate(color_map.items()):
        sample_lessons = set(cell[2] for cell in cells if cell[2] and not re.search(r'\\d{4}-\\d{2}-\\d{2}|MONDAY|TUESDAY|STUDIO', cell[2]))
        print(f"Color {i+1} ({color}): {len(cells)} cells")
        if sample_lessons:
            print(f"  Sample lessons: {list(sample_lessons)[:3]}")
    
    return teacher_df, lesson_df, color_map

if __name__ == "__main__":
    teacher_df, lesson_df, color_map = extract_teachers_by_color()
    
    print("\\nFirst few rows of teacher colors:")
    print(teacher_df.head().to_string())
    
    print("\\nFirst few rows of lesson data:")
    print(lesson_df.head().to_string())
