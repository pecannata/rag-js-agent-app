from openpyxl import load_workbook
import pandas as pd
import re

def extract_teachers_by_color_fixed():
    """Extract cell colors from the 69-615 sheet to identify teachers - FIXED VERSION"""
    
    file_path = 'Studio database/Private lesson Calendar.xlsx'
    sheet_name = '69 -615'
    
    # Load workbook with data_only=True to get calculated values
    wb_data = load_workbook(file_path, data_only=True)
    ws_data = wb_data[sheet_name]
    
    # Load workbook with data_only=False to get formatting
    wb_format = load_workbook(file_path, data_only=False)
    ws_format = wb_format[sheet_name]
    
    print(f"=== EXTRACTING TEACHERS BY COLOR: {sheet_name} ===")
    print(f"Sheet dimensions: {ws_data.max_row} rows × {ws_data.max_column} columns")
    
    # Get all cell data including colors
    cell_data = {}
    color_map = {}
    
    for row in range(1, ws_data.max_row + 1):
        for col in range(1, ws_data.max_column + 1):
            cell_data_val = ws_data.cell(row=row, column=col)
            cell_format_val = ws_format.cell(row=row, column=col)
            
            if cell_data_val.value is not None:
                # Get cell value (calculated if formula)
                value = str(cell_data_val.value).strip()
                
                # Get cell fill color from formatting workbook
                fill = cell_format_val.fill
                color = None
                
                if fill and fill.start_color and fill.start_color.rgb:
                    if fill.start_color.rgb != '00000000':  # Not default/no color
                        color = fill.start_color.rgb
                        
                cell_data[(row, col)] = {
                    'value': value,
                    'color': color
                }
                
                # Track unique colors
                if color and color not in color_map:
                    color_map[color] = []
                if color:
                    color_map[color].append((row, col, value))
    
    print(f"Found {len(cell_data)} cells with data")
    print(f"Found {len(color_map)} unique colors")
    
    # Analyze color patterns
    print("\\n=== COLOR ANALYSIS ===")
    for i, (color, cells) in enumerate(color_map.items()):
        print(f"Color {i+1} ({color}): {len(cells)} cells")
        # Show sample cells for this color (excluding headers/dates)
        lesson_samples = [cell[2] for cell in cells if cell[2] and not re.search(r'\\d{4}-\\d{2}-\\d{2}|DAILY|SCHEDULE|STUDIO|Week|MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY|\\d{2}:\\d{2}:\\d{2}', cell[2])]
        if lesson_samples:
            print(f"  Sample lessons: {lesson_samples[:3]}")
    
    # Find structure
    date_row = 3
    day_row = 4  
    studio_row = 5
    time_start_row = 6
    
    # Extract day-studio columns - handle both direct values and formulas
    day_studio_cols = []
    for col in range(1, ws_data.max_column + 1):
        day_val = cell_data.get((day_row, col), {}).get('value', '').upper()
        studio_val = cell_data.get((studio_row, col), {}).get('value', '')
        
        # Handle day names (including calculated ones)
        if day_val in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY'] and 'STUDIO' in studio_val.upper():
            day_studio_cols.append((col, day_val, studio_val))
    
    print(f"\\nFound {len(day_studio_cols)} day-studio combinations:")
    for col, day, studio in day_studio_cols:
        print(f"  Col {col}: {day} - {studio}")
    
    # Extract lesson data with colors (teachers)
    print("\\n=== EXTRACTING LESSONS WITH TEACHER COLORS ===")
    
    teacher_data = []
    lesson_data = []
    
    for row in range(time_start_row, ws_data.max_row + 1):
        # Get time from column 2
        time_cell = cell_data.get((row, 2), {})
        time_val = time_cell.get('value', '')
        
        if re.search(r'\d{1,2}:\d{2}:\d{2}', time_val):
            teacher_row = [time_val]
            lesson_row = [time_val]
            
            for col, day, studio in day_studio_cols:
                cell_info = cell_data.get((row, col), {})
                value = cell_info.get('value', '')
                color = cell_info.get('color', '')
                
                teacher_row.append(color if color else '')
                lesson_row.append(value)
            
            # Only add if there's actual lesson data
            if any(cell.strip() for cell in lesson_row[1:]):
                teacher_data.append(teacher_row)
                lesson_data.append(lesson_row)
                
                # Show color info for this time slot
                colored_cells = [color for color in teacher_row[1:] if color]
                if colored_cells:
                    print(f"Time {time_val}: {len(colored_cells)} colored cells")
    
    # Create headers
    headers = ['Time']
    for col, day, studio in day_studio_cols:
        # Get date for this column
        date_val = cell_data.get((date_row, col), {}).get('value', '')
        if '2025' in str(date_val):
            date = str(date_val).split()[0]
        else:
            # Try nearby columns for date
            for d_col in range(max(1, col-2), min(ws_data.max_column+1, col+3)):
                date_val = cell_data.get((date_row, d_col), {}).get('value', '')
                if '2025' in str(date_val):
                    date = str(date_val).split()[0]
                    break
            else:
                date = ''
        headers.append(f"{day} {studio} ({date})")
    
    # Save teacher colors CSV
    teacher_df = pd.DataFrame(teacher_data, columns=headers)
    teacher_output = 'Studio database/Private_lesson_Calendar_69-615_TEACHERS_BY_COLOR.csv'
    teacher_df.to_csv(teacher_output, index=False)
    
    # Save lesson data CSV  
    lesson_df = pd.DataFrame(lesson_data, columns=headers)
    lesson_output = 'Studio database/Private_lesson_Calendar_69-615_LESSONS_WITH_COLORS.csv'
    lesson_df.to_csv(lesson_output, index=False)
    
    print(f"\\n=== RESULTS ===")
    print(f"Teacher colors saved to: {teacher_output}")
    print(f"Lesson data saved to: {lesson_output}")
    print(f"Time slots: {len(teacher_data)}")
    
    # Create color mapping guide
    print("\\n=== COLOR TO TEACHER MAPPING GUIDE ===")
    for i, (color, cells) in enumerate(color_map.items()):
        lesson_samples = set(cell[2] for cell in cells if cell[2] and not re.search(r'\\d{4}-\\d{2}-\\d{2}|DAILY|SCHEDULE|STUDIO|Week|MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY|\\d{2}:\\d{2}:\\d{2}', cell[2]))
        if lesson_samples:
            print(f"Teacher {i+1} (Color {color}): {list(lesson_samples)[:5]}")
    
    return teacher_df, lesson_df, color_map

if __name__ == "__main__":
    teacher_df, lesson_df, color_map = extract_teachers_by_color_fixed()
    
    print("\\nFirst few rows of teacher colors:")
    print(teacher_df.head().to_string())
    
    print("\\nFirst few rows of lesson data:")
    print(lesson_df.head().to_string())
