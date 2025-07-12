from openpyxl import load_workbook
import pandas as pd
import re
import json
from datetime import datetime

def analyze_schedule_with_teachers():
    """Combine lesson data with teacher colors to create comprehensive schedule JSON"""
    
    file_path = 'Private lesson Calendar.xlsx'
    sheet_name = '69 -615'
    
    print(f"=== COMBINED SCHEDULE ANALYSIS: {sheet_name} ===")
    
    # Load workbook with data_only=True to get calculated values
    wb_data = load_workbook(file_path, data_only=True)
    ws_data = wb_data[sheet_name]
    
    # Load workbook with data_only=False to get formatting
    wb_format = load_workbook(file_path, data_only=False)
    ws_format = wb_format[sheet_name]
    
    print(f"Sheet dimensions: {ws_data.max_row} rows × {ws_data.max_column} columns")
    
    # Get all cell data including colors and values
    cell_data = {}
    color_map = {}
    
    for row in range(1, ws_data.max_row + 1):
        for col in range(1, ws_data.max_column + 1):
            cell_data_val = ws_data.cell(row=row, column=col)
            cell_format_val = ws_format.cell(row=row, column=col)
            
            if cell_data_val.value is not None:
                # Get cell value (calculated if formula)
                value = str(cell_data_val.value).strip()
                
                # Get cell fill color from formatting workbook
                fill = cell_format_val.fill
                color = None
                
                if fill and fill.start_color and fill.start_color.rgb:
                    if fill.start_color.rgb != '00000000':  # Not default/no color
                        color = fill.start_color.rgb
                        
                cell_data[(row, col)] = {
                    'value': value,
                    'color': color
                }
                
                # Track unique colors
                if color and color not in color_map:
                    color_map[color] = []
                if color:
                    color_map[color].append((row, col, value))
    
    print(f"Found {len(cell_data)} cells with data")
    print(f"Found {len(color_map)} unique colors")
    
    # Identify teacher names from colors
    print("\\n=== IDENTIFYING TEACHERS FROM COLORS ===")
    teacher_colors = {}
    
    for color, cells in color_map.items():
        # Look for explicit teacher names (ALL CAPS entries that look like names)
        teacher_names = []
        for row, col, value in cells:
            # Check if this looks like a teacher name
            if (value.isupper() and 
                len(value) > 2 and 
                not re.search(r'\\d{4}-\\d{2}-\\d{2}|DAILY|SCHEDULE|STUDIO|MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY|\\d{2}:\\d{2}', value) and
                any(name in value for name in ['RYANN', 'GRACIE', 'CARALIN', 'HUNTER', 'ARDEN', 'MEGHAN', 'PAIGE'])):
                teacher_names.append(value)
        
        if teacher_names:
            teacher_colors[color] = teacher_names[0]  # Take first found name
        else:
            # For colors without explicit teacher names, treat as unassigned
            teacher_colors[color] = None
    
    print("Teacher-Color Mapping:")
    for color, teacher in teacher_colors.items():
        print(f"  {teacher}: {color}")
    
    # Find structure
    date_row = 3
    day_row = 4
    studio_row = 5
    time_start_row = 6
    
    # Extract dates
    dates = []
    for col in range(1, ws_data.max_column + 1):
        date_val = cell_data.get((date_row, col), {}).get('value', '')
        if '2025' in str(date_val):
            dates.append(str(date_val).split()[0])
    
    # Extract day-studio combinations
    day_studio_cols = []
    for col in range(1, ws_data.max_column + 1):
        day_val = cell_data.get((day_row, col), {}).get('value', '').upper()
        studio_val = cell_data.get((studio_row, col), {}).get('value', '')
        
        if day_val in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY'] and 'STUDIO' in studio_val.upper():
            # Find corresponding date
            date = ''
            for d_col in range(max(1, col-2), min(ws_data.max_column+1, col+3)):
                date_val = cell_data.get((date_row, d_col), {}).get('value', '')
                if '2025' in str(date_val):
                    date = str(date_val).split()[0]
                    break
            
            day_studio_cols.append({
                'column': col,
                'day': day_val,
                'studio': studio_val,
                'date': date
            })
    
    print(f"\\nFound {len(day_studio_cols)} day-studio combinations")
    
    # Extract complete schedule data
    print("\\n=== EXTRACTING COMPLETE SCHEDULE ===")
    
    schedule_data = {
        'week_info': {
            'sheet_name': sheet_name,
            'week_dates': dates,
            'extracted_date': datetime.now().isoformat()
        },
        'teachers': teacher_colors,
        'studios': list(set(item['studio'] for item in day_studio_cols)),
        'schedule': []
    }
    
    # Process each time slot
    for row in range(time_start_row, ws_data.max_row + 1):
        time_cell = cell_data.get((row, 2), {})
        time_val = time_cell.get('value', '')
        
        if re.search(r'\d{1,2}:\d{2}:\d{2}', time_val):
            time_slot = {
                'time': time_val,
                'lessons': []
            }
            
            for studio_info in day_studio_cols:
                col = studio_info['column']
                cell_info = cell_data.get((row, col), {})
                value = cell_info.get('value', '')
                color = cell_info.get('color', '')
                
                if value.strip():  # Only include if there's lesson data
                    lesson = {
                        'day': studio_info['day'],
                        'date': studio_info['date'],
                        'studio': studio_info['studio'],
                        'student_info': value,
                        'teacher_color': color,
                        'teacher': teacher_colors.get(color, 'Unknown'),
                        'column': col
                    }
                    time_slot['lessons'].append(lesson)
            
            # Only add time slot if it has lessons
            if time_slot['lessons']:
                schedule_data['schedule'].append(time_slot)
    
    print(f"Extracted {len(schedule_data['schedule'])} time slots")
    
    # Add summary statistics
    schedule_data['summary'] = {
        'total_time_slots': len(schedule_data['schedule']),
        'total_lessons': sum(len(slot['lessons']) for slot in schedule_data['schedule']),
        'unique_teachers': len(set(teacher_colors.values())),
        'date_range': f"{min(dates)} to {max(dates)}" if dates else "Unknown"
    }
    
    return schedule_data

def save_schedule_json():
    """Extract schedule data and save as JSON"""
    
    schedule_data = analyze_schedule_with_teachers()
    
    # Save to JSON file
    output_file = 'Private_lesson_Calendar_69-615_COMPLETE_SCHEDULE.json'
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(schedule_data, f, indent=2, ensure_ascii=False)
    
    print(f"\\n=== RESULTS ===")
    print(f"Complete schedule saved to: {output_file}")
    print(f"\\nSummary:")
    print(f"  Week: {schedule_data['summary']['date_range']}")
    print(f"  Time slots: {schedule_data['summary']['total_time_slots']}")
    print(f"  Total lessons: {schedule_data['summary']['total_lessons']}")
    print(f"  Teachers: {schedule_data['summary']['unique_teachers']}")
    print(f"  Studios: {len(schedule_data['studios'])}")
    
    # Show sample data
    print(f"\\nSample lesson data:")
    if schedule_data['schedule']:
        first_slot = schedule_data['schedule'][0]
        print(f"  Time: {first_slot['time']}")
        for lesson in first_slot['lessons'][:3]:  # Show first 3 lessons
            print(f"    {lesson['day']} {lesson['studio']}: {lesson['student_info']} (Teacher: {lesson['teacher']})")
    
    return schedule_data

if __name__ == "__main__":
    schedule_data = save_schedule_json()
    
    # Also create a simplified CSV summary
    print("\\n=== CREATING SIMPLIFIED SUMMARY ===")
    
    lessons_list = []
    for time_slot in schedule_data['schedule']:
        for lesson in time_slot['lessons']:
            lessons_list.append({
                'Time': time_slot['time'],
                'Day': lesson['day'],
                'Date': lesson['date'],
                'Studio': lesson['studio'],
                'Student_Info': lesson['student_info'],
                'Teacher': lesson['teacher'],
                'Teacher_Color': lesson['teacher_color']
            })
    
    df = pd.DataFrame(lessons_list)
    summary_file = 'Private_lesson_Calendar_69-615_LESSON_SUMMARY.csv'
    df.to_csv(summary_file, index=False)
    
    print(f"Lesson summary saved to: {summary_file}")
    print(f"Total records: {len(lessons_list)}")
