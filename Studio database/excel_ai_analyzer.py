from openpyxl import load_workbook
import csv
import re
from datetime import datetime

def analyze_excel_with_ai_logic(file_path, sheet_name):
    """
    Use openpyxl to get raw cell values and apply AI-like logic
    to understand the structure and extract meaningful data
    """
    
    # Load workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    print(f"=== AI-POWERED EXCEL ANALYSIS: {sheet_name} ===")
    print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
    print()
    
    # Step 1: Find structural elements
    print("=== IDENTIFYING STRUCTURE ===")
    
    header_row = None
    date_row = None
    day_row = None
    time_slots = []
    
    # Scan all cells to understand structure
    for row in range(1, min(15, ws.max_row + 1)):  # Check first 15 rows
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                row_data.append(str(cell.value).strip())
        
        if not row_data:
            continue
            
        # Check if this row contains dates
        date_count = sum(1 for item in row_data if re.search(r'2024-\d{2}-\d{2}', item))
        if date_count >= 3:
            date_row = row
            print(f"Row {row}: DATE ROW - {row_data}")
            continue
            
        # Check if this row contains day names
        day_names = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        day_count = sum(1 for item in row_data if item.upper() in day_names)
        if day_count >= 3:
            day_row = row
            print(f"Row {row}: DAY ROW - {row_data}")
            continue
            
        # Check if this row contains time slots
        has_time = any(re.search(r'\d{1,2}:\d{2}', item) for item in row_data)
        if has_time:
            time_slots.append(row)
            print(f"Row {row}: TIME SLOT - {row_data}")
    
    print(f"\\nFound: Date row={date_row}, Day row={day_row}, Time slots={time_slots}")
    
    # Step 2: Extract dates and days
    dates = []
    days = []
    
    if date_row:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=date_row, column=col)
            if cell.value and '2024' in str(cell.value):
                dates.append(str(cell.value).split()[0])
    
    if day_row:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=day_row, column=col)
            if cell.value:
                day = str(cell.value).strip().upper()
                if day in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
                    days.append(day)
    
    print(f"\\nExtracted dates: {dates}")
    print(f"Extracted days: {days}")
    
    # Step 3: Extract ALL lesson data with complete coverage
    print("\\n=== EXTRACTING COMPLETE LESSON DATA ===")
    
    csv_data = []
    headers = ['Time'] + [f"{day} ({date})" for day, date in zip(days, dates)]
    csv_data.append(headers)
    
    for time_row in time_slots:
        row_data = []
        
        # Get the time from the first column that contains a time
        time_value = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=time_row, column=col)
            if cell.value:
                cell_str = str(cell.value)
                time_match = re.search(r'\d{1,2}:\d{2}(?::\d{2})?(?:\s*(?:AM|PM))?', cell_str, re.IGNORECASE)
                if time_match and not time_value:
                    time_value = time_match.group()
                    break
        
        if not time_value:
            continue
            
        row_data.append(time_value)
        
        # Extract lesson data for each day column
        # Find which columns correspond to days
        day_columns = []
        if day_row:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=day_row, column=col)
                if cell.value:
                    day = str(cell.value).strip().upper()
                    if day in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
                        day_columns.append(col)
        
        # Get lesson data for each day
        for col in day_columns:
            cell = ws.cell(row=time_row, column=col)
            if cell.value:
                cell_str = str(cell.value).strip()
                # Clean up the cell content
                if cell_str and cell_str != 'nan':
                    # Remove redundant time if it's the same as the row time
                    if time_value.replace(':00', '') not in cell_str:
                        row_data.append(cell_str)
                    else:
                        # Extract just the non-time part
                        cleaned = re.sub(r'\d{1,2}:\d{2}(?::\d{2})?(?:\s*(?:AM|PM))?:?\s*', '', cell_str, flags=re.IGNORECASE)
                        row_data.append(cleaned.strip() if cleaned.strip() else cell_str)
                else:
                    row_data.append("")
            else:
                row_data.append("")
        
        csv_data.append(row_data)
        print(f"Time {time_value}: {row_data[1:]}")
    
    return csv_data

def save_to_csv(data, filename):
    """Save the extracted data to CSV"""
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        for row in data:
            writer.writerow(row)
    
    print(f"\\nSaved to: {filename}")

# Run the analysis
if __name__ == "__main__":
    excel_file = 'Private lesson Calendar.xlsx'
    sheet_name = '99-915'
    output_file = 'Private_lesson_Calendar_tab2_99-915_AI_COMPLETE.csv'
    
    try:
        csv_data = analyze_excel_with_ai_logic(excel_file, sheet_name)
        save_to_csv(csv_data, output_file)
        
        print(f"\\n=== FINAL RESULT ===")
        print(f"Extracted {len(csv_data)-1} time slots")
        print(f"CSV saved as: {output_file}")
        
        # Show the final result
        import pandas as pd
        df = pd.read_csv(output_file)
        print("\\nFinal CSV content:")
        print(df.to_string())
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
