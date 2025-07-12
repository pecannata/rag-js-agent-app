from openpyxl import load_workbook
import pandas as pd
import re

def analyze_69_615_sheet():
    """Complete analysis of the 69-615 sheet"""
    
    file_path = 'Private_lesson_Calendar.xlsx'
    sheet_name = '69 -615'  # Note the space
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    print(f"=== CLAUDE ANALYSIS: {sheet_name} ===")
    print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Get all non-empty cells
    all_data = {}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                all_data[(row, col)] = str(cell.value).strip()
    
    print(f"Found {len(all_data)} non-empty cells")
    
    # Identify structure - this appears to be a multi-studio schedule
    print("\n=== STRUCTURE IDENTIFICATION ===")
    
    # Find key rows
    date_row = None
    day_row = None
    studio_row = None
    time_start_row = None
    
    for row in range(1, 10):
        row_values = [all_data.get((row, col), '') for col in range(1, ws.max_column + 1)]
        
        # Check for dates
        date_count = sum(1 for val in row_values if re.search(r'2025-\d{2}-\d{2}', val))
        if date_count >= 5:
            date_row = row
            print(f"Date row: {row}")
        
        # Check for day names
        day_names = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        day_count = sum(1 for val in row_values if val.upper() in day_names)
        if day_count >= 5:
            day_row = row
            print(f"Day row: {row}")
        
        # Check for studios
        studio_count = sum(1 for val in row_values if 'STUDIO' in val.upper())
        if studio_count >= 3:
            studio_row = row
            print(f"Studio row: {row}")
        
        # Check for times
        time_count = sum(1 for val in row_values if re.search(r'\d{1,2}:\d{2}:\d{2}', val))
        if time_count >= 1 and not time_start_row:
            time_start_row = row
            print(f"Time data starts: {row}")
    
    # Extract structure information
    print(f"\nStructure: Date={date_row}, Day={day_row}, Studio={studio_row}, Times={time_start_row}")
    
    # Extract dates
    dates = []
    if date_row:
        for col in range(1, ws.max_column + 1):
            val = all_data.get((date_row, col), '')
            if '2025' in val:
                dates.append(val.split()[0])
    
    # Extract days and studios
    day_studio_cols = []
    if day_row and studio_row:
        for col in range(1, ws.max_column + 1):
            day = all_data.get((day_row, col), '').upper()
            studio = all_data.get((studio_row, col), '')
            if day in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
                day_studio_cols.append((col, day, studio))
    
    print(f"\nFound {len(dates)} dates: {dates[:5]}...")  # Show first 5
    print(f"Found {len(day_studio_cols)} day-studio combinations:")
    for col, day, studio in day_studio_cols[:10]:  # Show first 10
        print(f"  Col {col}: {day} - {studio}")
    
    # Extract lesson data
    print(f"\n=== EXTRACTING LESSON DATA ===")
    
    lesson_data = []
    
    # Process time rows
    for row in range(time_start_row or 6, ws.max_row + 1):
        # Get time from column 2 (where times seem to be)
        time_val = all_data.get((row, 2), '')
        
        if re.search(r'\d{1,2}:\d{2}:\d{2}', time_val):
            row_data = [time_val]
            
            # Get lesson data for each day-studio column
            for col, day, studio in day_studio_cols:
                cell_val = all_data.get((row, col), '')
                if cell_val:
                    # Format as "STUDENT (if any)"
                    row_data.append(cell_val)
                else:
                    row_data.append('')
            
            # Only add if there's actual lesson data
            if any(cell.strip() for cell in row_data[1:]):
                lesson_data.append(row_data)
                print(f"Time {time_val}: {len([x for x in row_data[1:] if x.strip()])} lessons")
    
    # Create headers
    headers = ['Time']
    for col, day, studio in day_studio_cols:
        # Get corresponding date
        date = ''
        if date_row:
            for d_col in range(col-2, col+3):  # Check nearby columns for date
                date_val = all_data.get((date_row, d_col), '')
                if '2025' in date_val:
                    date = date_val.split()[0]
                    break
        
        if not date and dates:
            # Try to match based on day order
            day_order = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
            if day in day_order and len(dates) > day_order.index(day):
                date = dates[day_order.index(day)]
        
        header = f"{day} {studio}\\n({date})" if date else f"{day} {studio}"
        headers.append(header)
    
    return headers, lesson_data

# Run the analysis
if __name__ == "__main__":
    headers, lesson_data = analyze_69_615_sheet()
    
    print(f"\n=== FINAL RESULTS ===")
    print(f"Headers ({len(headers)}): {headers[:5]}...")  # Show first 5
    print(f"Lesson rows: {len(lesson_data)}")
    
    # Create and save CSV
    df = pd.DataFrame(lesson_data, columns=headers)
    output_file = 'Private_lesson_Calendar_69-615_COMPLETE.csv'
    df.to_csv(output_file, index=False)
    
    print(f"\nSaved to: {output_file}")
    print("\nFirst few rows:")
    print(df.head(10).to_string())
    
    # Show summary
    print(f"\n=== SUMMARY ===")
    print(f"Time slots: {len(lesson_data)}")
    print(f"Studios/columns: {len(headers)-1}")
    
    # Count lessons
    total_lessons = sum(1 for row in lesson_data for cell in row[1:] if cell.strip())
    print(f"Total lessons scheduled: {total_lessons}")
